'''
Program nariše zemljevid na podlagi meritev iz geodetskih merjenj.

Install required libraries:
Windows: py -m pip install openpyxl matplotlib numpy
Unix/macOS: python -m pip install openpyxl matplotlib numpy
'''

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import math
import os

# Ustvari datoteko "koordinati.xlsx" in v njeno prvo vrstico zapiše spodnje označbe.
koordinati1 = openpyxl.Workbook()
ws2 = koordinati1.active
ws2.cell(column=1, row=1).value = "Stojišče"
ws2.cell(column=2, row=1).value = "Točka"
ws2.cell(column=3, row=1).value = "x"
ws2.cell(column=4, row=1).value = "y"
ws2.cell(column=5, row=1).value = "Px"
ws2.cell(column=6, row=1).value = "Py"

# V datoteki "koordinati.xslx" funkcija pregleda izbran stolpec, v njem išče celico z vrednostjo None in vrne številko njene vrstice. To je zadnja prazna, ne polna vrstica; zato "erow" ("empty row"). (!) Predpostavljeno je, da stolpec ne vsebuje vmesnih None vrednosti.
def last_erow(choosen_column):
	last_erow = 0
	cell = True
	while cell:
		last_erow += 1
		cell = ws2.cell(column=choosen_column, row=last_erow).value
	return last_erow

# Funkcija razcepi seznam v seznam s podseznami glede na ločevalca (ang. delimiter). Npr. [1, 2, 4, "E", 5, 3, "E"] gre v [[1, 2, 4], [5, 3]]
def split_list(input_list, delimiter):
	output_list = []
	inner_list = []
	for i in input_list:
		if i == delimiter:
			output_list += [inner_list]
			inner_list = []
			continue
		inner_list += [i]
	return output_list

# Prebere direktorij "meritve/" in za vsako datoteko preveri, ali se njeno ime začne s "." ali "i". Slednje datoteke izbriše (iz seznama). Navadna for-zanka tukaj ne deluje pravilno, ker se spreminjajo indexi seznama.
meritve = os.listdir("meritve/")
i = 0
while i != len(meritve):
	meritev = list(meritve[i])
	if "." in meritev[0] or "i" in meritev[0]:
		del meritve[i]
		i = i - 1
	i = i + 1
meritve.sort()

# Zanka bere meriteve in zapisuje izračune v posamezne datoteke ter v "koordinati.xlsx". 
for datoteka in meritve:
	filename = "meritve/" + datoteka
	wb = openpyxl.load_workbook(filename)
	ws1 = wb['Sheet1']

	stojiščE = ws1.cell(column=1, row=2).value
	orientacijA = ws1.cell(column=1, row=5).value

	# Seznami, v katere se bodo shranjevali izračuni.
	spremembe = []
	koordinati = []
	povezave = []

	# Prebere azimut in ga pretvori v radiane. Če je azimut 0, mora to v celici pisati in ne sme biti prazna!
	azimutS = float(ws1.cell(column=1, row=8).value)
	azimutM = float(ws1.cell(column=2, row=8).value)
	azimut = (azimutS + azimutM/60) * math.pi/180
	
# IZRAČUNI -----------------------------------------------------------------------

	# Izračuna zdajšnja kota alfa in beta, pri prvem stojišču malo drugače. Kot alfa je kot med premico, ki poteka skozi zdajšnje in prejšnje stojišče, ter premico, ki poteka skozi zdajšnje in naslednje stojišče (orientacijo). Kot betaZ je komplementarnen kotu alphaZ.
	if datoteka == "1001.xlsx":
		alphaP = math.pi/2
	alphaZ = abs(azimut - alphaP)
	betaZ = math.pi/2 - alphaZ

	# x (abscisa) orientacje
	xo = float(ws1.cell(column=6, row=2).value)

	# Zanka prebere vrstico iz meritev ter opravi izračune. Upravljanje z izjemami sem uporabil zaradi čudnega obnašanja vrednosti None: Na primer, če v preglednici zbrišemo zadnje tri vrednosti, jih python zazna kot tipa NoneType in vrne napako pri računanju.
	for i in range(2, ws1.max_row + 1):
		try:
			točka = int(ws1.cell(column=3, row=i).value)
			kotS = float(ws1.cell(column=4, row=i).value)
			kotM = float(ws1.cell(column=5, row=i).value)
			d = float(ws1.cell(column=6, row=i).value) # dolžina
			povezava = ws1.cell(column=8, row=i).value
		except:
			print("NoneType found")
			break
		
		# Formule za izračun točke v posameznem koordinatnem sistemu.
		kot = (kotS + kotM/60) * math.pi/180
		deltaX = math.cos(kot) * d
		deltaY = math.sin(kot) * d
		
		# Formule za izračun točke v skupnem koordinatnem sistemu.
		m = deltaX - deltaY * math.tan(betaZ)
		X = xo + m * math.cos(alphaZ) + deltaY/math.cos(betaZ)
		Y = m * math.sin(alphaZ)
		
		spremembe += [[točka, deltaX, deltaY]] # npr. [[6, 5.67, 3.45], [7, 2.33, 6.9],...]
		koordinati += [[X, Y]]
		
		# Povezavo, ki je niz (ang. string), razcepi z ločavalcem (ang. delimiter) "-" in jo s tem spremeni v seznam točk, ki bodo kasneje povezane. Seznamu povezave doda podseznam povezava.
		if povezava is not None:
			povezava = povezava.split("-")
			povezave += [povezava]
		
		#print("Točka: " + str(točka), "Kot: " + str(kot), "Dolžina: " + str(d), "Δx: " + str(deltaX), "Δy: " + str(deltaY), "", sep = "\n")

	# Ker je konec izračunov, prejšnji alfa dobi vrednost zdajšnjega.
	alphaP = alphaZ
	
# ZAPIS V POSAMEZNE DATOTEKE -----------------------------------
	
	# Ustvari datotekom za izračune, npr. "meritve/i1001.xlsx"
	izračuni = openpyxl.Workbook()
	dest_filename = "meritve/" + "i" + datoteka
	ws1 = izračuni.active

	# Zapiše oznake.
	ws1.cell(column=1, row=1).value = "Stojišče"
	ws1.cell(column=1, row=4).value = "Orientacija"
	ws1.cell(column=1, row=7).value = "Azimut[°]"
	ws1.cell(column=2, row=7).value = "Azimut[']"
	ws1.cell(column=3, row=1).value = "Točka"
	ws1.cell(column=4, row=1).value = "Δx"
	ws1.cell(column=5, row=1).value = "Δy"
	
	ws1.cell(column=1, row=2).value = stojiščE
	ws1.cell(column=1, row=5).value = orientacijA
	ws1.cell(column=1, row=8).value = azimutS
	ws1.cell(column=2, row=8).value = azimutM

	# Zapiše izračune.
	for i in range(len(spremembe)):
		točka = spremembe[i][0]
		deltaX = spremembe[i][1]
		deltaY = spremembe[i][2]
		
		ws1.cell(column=3, row=i + 2).value = točka
		ws1.cell(column=4, row=i + 2).value = deltaX
		ws1.cell(column=5, row=i + 2).value = deltaY

	izračuni.save(filename = dest_filename)
	
# ZAPIS V SKUPNO DATOTEKO -------------------------------------------

	# Shrani trenutno zadnjo vrstico, kajti kasneje se na to vrednost zanaša (*) in se ne sme preračunavati vsakič na novo.
	lrow = last_erow(2)
	ws2.cell(column=1, row=lrow).value = stojiščE
	
	# Zapiše izračune v "koordinati.xslx".
	for i in range(len(spremembe)):
		točka = spremembe[i][0]
		X = koordinati[i][0]
		Y = koordinati[i][1]
		
		i += lrow # (*)
		ws2.cell(column=2, row=i).value = točka
		ws2.cell(column=3, row=i).value = X
		ws2.cell(column=4, row=i).value = Y
		
	# Zapiše povezave.
	# Zanka gre skozi seznam povezave.
	for i in range(len(povezave)):
		#x = []
		#y = []
		
		# Zanka gre skozi elemente posameznega seznama v seznamu povezave.
		for j in range(len(povezave[i])):
		
			# Vzame posamezno točko ene povezave.
			točka = povezave[i][j]
			Px = koordinati[int(točka)][0]
			Py = koordinati[int(točka)][1]
			
			# Zapiše to točko v "koordinati.xlsx".
			c = 5
			ws2.cell(column=c, row=last_erow(c)).value = Px
			ws2.cell(column=c+1, row=last_erow(c+1)).value = Py
		
		# Konec te povezave označi z "E" ('end').
		delimiter = "E"
		ws2.cell(column=c, row=last_erow(c)).value = delimiter
		ws2.cell(column=c+1, row=last_erow(c+1)).value = delimiter
		
		#plt.plot(x, y, '-', color="#ff8c00", zorder=1)
	
koordinati1.save(filename = "koordinati.xlsx")

# GRAF ----------------------------------------------

# Prebere "koordinati.xlsx".
x = []
y = []
Px = []
Py = []
for i in range(2, last_erow(3)):
	xpoint = ws2.cell(column=3, row=i).value
	ypoint = ws2.cell(column=4, row=i).value
	Pxpoint = ws2.cell(column=5, row=i).value
	Pypoint = ws2.cell(column=6, row=i).value
	x += [xpoint]
	y += [ypoint]
	Px += [Pxpoint]
	Py += [Pypoint]

# Razcepi seznam absic in ordinat povezav.
Px = split_list(Px, delimiter)
Py = split_list(Py, delimiter)

# Nastavitve koordinatnega sistema.
fig, ax = plt.subplots()

ax.spines['right'].set_visible(False)
ax.spines['left'].set_visible(False)
ax.spines['top'].set_visible(False)
ax.spines['bottom'].set_visible(False)

plt.xticks([])
plt.yticks([])

#max_value = max([max(x), max(y)])
#plt.xlim(0, max_value + 1)
#plt.ylim(0, max_value + 1)

# Označi vse točke.
for i in range(len(x)):
	plt.text(x[i], y[i] + 0.1, i+1, fontsize="xx-small")

# Nariše posamezne povezave in nato točke.
for i in range(len(Px)):
	plt.plot(np.array(Px[i]), np.array(Py[i]), '-', color="#000000")
plt.plot(np.array(x), np.array(y), '.', color="#ffa500")

plt.show()


