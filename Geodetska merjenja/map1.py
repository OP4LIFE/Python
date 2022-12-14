#------------ README ----------------------------
'''
The program draws a map from given coordinates.

Install required libraries:
Windows: py -m pip install openpyxl matplotlib numpy
Unix/macOS: python -m pip install openpyxl matplotlib numpy
'''

# Add functions, i confition, checker

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Load a file.
file = 'koordinati1.xlsx'
wb = openpyxl.load_workbook(file)


#------------ READ ----------------------------
sheet1 = wb['Sheet1']

xpoints = []
ypoints = []
for i in range(1, sheet1.max_row + 1):
    
    # Read values at current row.
    x = sheet1.cell(column=1, row=i).value
    y = sheet1.cell(column=2, row=i).value
    
    # Apply the formula.
    x += 1
    y += 1
    
    # Append points to the lists.
    xpoints.append(x)
    ypoints.append(y)

#------------ WRITE ----------------------------
# Create Sheet2.
if 'Sheet2' not in wb.sheetnames:
    wb.create_sheet('Sheet2')
sheet2 = wb['Sheet2']

# Write to Sheet2.
for i in range(1, sheet1.max_row + 1):
    sheet2.cell(column=1, row=i).value = xpoints[i - 1]
    sheet2.cell(column=2, row=i).value = ypoints[i - 1]

wb.save(file)


#------------ DRAW ----------------------------
# Convert lists to arrays.
xpoints = np.array(xpoints)
ypoints = np.array(ypoints)

# Axes' settings.
fig, ax = plt.subplots()

ax.set_aspect('equal')

ax.spines['right'].set_visible(False)
ax.spines['left'].set_visible(False)
ax.spines['top'].set_visible(False)
ax.spines['bottom'].set_visible(False)

plt.xticks([])
plt.yticks([])

max_value = max([max(xpoints), max(ypoints)])
plt.xlim(0, max_value + 1)
plt.ylim(0, max_value + 1)

# Label each point.
for i in range(1, sheet1.max_row):
    plt.text(xpoints[i - 1], ypoints[i - 1] + 0.2, str(i))

# Draw and save the map.
plt.plot(xpoints, ypoints, '.-')
plt.savefig('Zemljevid.svg')
plt.show()

