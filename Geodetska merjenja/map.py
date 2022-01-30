
#------------ README ----------------------------
'''
Install required libraries from the Command prompt or Terminal.
Windows: py -m pip install openpyxl matplotlib numpy
Unix/macOS: python -m pip install openpyxl matplotlib numpy
'''

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Load a file.
file = 'koordinati.xlsx'
wb = openpyxl.load_workbook(file)


#------------ READ ----------------------------
sheet1 = wb['Sheet1']

xpoints = []
ypoints = []
for i in range(1, sheet1.max_row + 1):
    
    # Read values at current row.
    x = sheet1.cell(column=1, row=i).value
    y = sheet1.cell(column=2, row=i).value
    
    # Apply the formula.
    x += 1
    y += 1
    
    # Append points to the lists.
    xpoints.append(x)
    ypoints.append(y)


#------------ WRITE ----------------------------
# Create Sheet2.
if 'Sheet2' not in wb.sheetnames:
    wb.create_sheet('Sheet2')
sheet2 = wb['Sheet2']

# Write to Sheet2.
for i in range(1, sheet1.max_row + 1):
    sheet2.cell(column=1, row=i).value = xpoints[i - 1]
    sheet2.cell(column=2, row=i).value = ypoints[i - 1]

wb.save(file)


#------------ DRAW ----------------------------
# Convert lists to arrays.
xpoints = np.array(xpoints)
ypoints = np.array(ypoints)

# Axes' settings.
fig, ax = plt.subplots()

ax.set_aspect('equal')

ax.spines['right'].set_visible(True)
ax.spines['top'].set_visible(True)

plt.xticks([])
plt.yticks([])

plt.xlim(0, 10)
plt.ylim(0, 10)

# Draw and save the map.
plt.plot(xpoints, ypoints, '.-')
plt.savefig('Zemljevid.svg')
plt.show()

