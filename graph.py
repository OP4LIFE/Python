#------------ README ----------------------------
'''
The program draws a map from given coordinates.

Install required libraries:
Windows: py -m pip install openpyxl matplotlib numpy
Unix/macOS: python -m pip install openpyxl matplotlib numpy
'''

import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Load a file.
file = 'koordinati.xlsx'
wb = openpyxl.load_workbook(file)


#------------ READ ----------------------------
sheet1 = wb['Sheet1']

xpoints = []
ypoints = []
for i in range(1, sheet1.max_row + 1):
    
    # Read values at current row.
    x = sheet1.cell(column=1, row=i).value
    y = sheet1.cell(column=2, row=i).value
    
    # Apply the formula.
#    x += 1
#    y += 1
    
    # Append points to the lists.
    xpoints.append(x)
    ypoints.append(y)

#------------ WRITE ----------------------------
# Create Sheet2.
if 'Sheet2' not in wb.sheetnames:
    wb.create_sheet('Sheet2')
sheet2 = wb['Sheet2']

# Write to Sheet2.
for i in range(1, sheet1.max_row + 1):
    sheet2.cell(column=1, row=i).value = xpoints[i - 1]
    sheet2.cell(column=2, row=i).value = ypoints[i - 1]

wb.save(file)


#------------ DRAW ----------------------------

# Convert lists to arrays.
xpoints = np.array(xpoints)
ypoints = np.array(ypoints)

# Similar curve
[1/4, 1/3, 1/2, 1, 2, 3, 4]
[16, 9, 4, 1, 1/4, 1/9, 1/16]

xcurve = 

# Axes' settings.
fig, ax = plt.subplots()

# Simetrical exes
#ax.set_aspect('equal')

ax.spines['right'].set_visible(True)
ax.spines['left'].set_visible(True)
ax.spines['top'].set_visible(True)
ax.spines['bottom'].set_visible(True)



# Ticks
padding = 1.07
ax.set_xticks(np.arange(0, max(xpoints) * padding, 5))
ax.set_xticks(np.arange(0, max(xpoints) * padding, 5/5), minor=True)
ax.set_yticks(np.arange(0, max(ypoints) * padding, 200))
ax.set_yticks(np.arange(0, max(ypoints) * padding, 200/5), minor=True)

# Enables (0, 0) to be on the axes' intersection.
plt.xlim(0, None)
plt.ylim(0, None)

# Grid
ax.grid(which='both')

# Grid transperency
ax.grid(which='minor', alpha=0.2)
ax.grid(which='major', alpha=0.5)

# Title and lables
plt.title("o(r)")
plt.xlabel('r[cm]', loc = "right")
plt.ylabel('o[lux]', loc = "top", rotation = 0)

# Draw and save the map.
plt.plot(xpoints, ypoints, '.')

#plt.savefig('Zemljevid.svg')
plt.show()


