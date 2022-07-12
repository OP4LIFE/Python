#------------ README ----------------------------
'''

Install required libraries:
Windows: py -m pip install openpyxl matplotlib numpy
Unix/macOS: python -m pip install openpyxl matplotlib numpy
'''

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from math import log10, floor

# Load a file.
file = 'graph.xlsx'
wb = openpyxl.load_workbook(file)


#------------ READ ----------------------------
sheet1 = wb['Sheet1']

x0 = []
y0 = []
for i in range(1, sheet1.max_row + 1):
    
    # Read values at current row.
    xpoints = sheet1.cell(column=1, row=i).value
    ypoints = sheet1.cell(column=2, row=i).value

    # Append points to the lists.
    x0.append(xpoints)
    y0.append(ypoints)

#------------ WRITE ----------------------------
# Create Sheet2.
if 'Sheet2' not in wb.sheetnames:
    wb.create_sheet('Sheet2')
sheet2 = wb['Sheet2']

# Write to Sheet2.
for i in range(1, sheet1.max_row + 1):
    sheet2.cell(column=1, row=i).value = x0[i - 1]
    sheet2.cell(column=2, row=i).value = y0[i - 1]

wb.save(file)


#------------ DRAW ----------------------------

def round1(value):
    return round(value, -int(floor(log10(abs(value)))))


def roundPartial (value, resolution):
    if value > 0 and value < 1:
         decimal_places = len(str(value).split(".")[1])
         resolution = resolution * 10 ** -decimal_places
         
    return round (value /float(resolution)) * resolution
    
    
'''
# Each point, calculates distance to line, returns average
def average_distance(slope, xpoints, ypoints):
    distances = []
    for i in range(len(x1)):
        x_ = xpoints[i]
        y_ = ypoints[i]
        
        d = abs(slope * x_ - y_) / (1 + slope ** 2) ** (1/2)       
        distances += [d]
               
    return sum(distances) / len(distances)  



def slope(xpoints, ypoints):
    k = [0.0000001, 1, 100000]
    for o in range(50):
          
    # Calculate averages.
        a = []  
        for p in range(len(k)):       
            a += [average_distance(k[p], xpoints, ypoints)]
        
    # Numerical searching for optimal slope.
        m = a.index(min(a))
        if m == 0:
            k = [k[m], (k[m] + k[m + 1])/2, k[m + 1]]
            
        elif m == len(a) - 1:
            k = [k[m - 1], (k[m - 1] + k[m])/2, k[m]]
            
        else:   
            k = [k[m - 1],    (k[m - 1] + k[m])/2,    k[m],    (k[m] + k[m + 1])/2,    k[m + 1]]
    #            previous            avarage         current       avarage               next    
        
        if o == 49:
            print("Coefficient: ", k[m])
            print("Average deviation: ", a[m])
            return int(k[m])
'''

def coefficient(xpoints, ypoints):
    k = [1000000, 1, 0.000001]
    iterations = 80
    finale = []
    for u in range(iterations):
        
        a = []
        for i in range(len(k)):      
            # Calculate mathematically perfect ordinates.
            accurate_y = []
            for x in xpoints:
                accurate_y.append(k[i] * x)

            # Calculate deviations of mesaured ordinates.
            deviations = []
            for o in range(len(ypoints)):
                deviations += [abs(ypoints[o] - accurate_y[o])]

            # Average deviation.
            a.append(sum(deviations) / len(deviations))

        # Numerical searching for optimal coefficient.
        m = a.index(min(a))
        if m == 0:
            k = [k[m], (k[m] + k[m + 1])/2, k[m + 1]]
                     
        elif m == len(a) - 1:
            k = [k[m - 1], (k[m - 1] + k[m])/2, k[m]] 
                
        else:   
            k = [k[m - 1],    (k[m - 1] + k[m])/2,    k[m],    (k[m] + k[m + 1])/2,    k[m + 1]]
    #            previous            avarage         current       avarage               next    
        
                
        if u == iterations - 1:
            print("\nCoefficient: ", int(k[m]))
            print("Average deviation: ", a[m], "\n")
            return int(k[m])
            
              
  
padding = 1.07
fig, ax = plt.subplots(2, 1)
k = 1

# Na začetku je linearizacija, da lahko najde optimalen smerni koeficient in ga nato uporabi v prvem grafu.
############# PLOT 2 (linearizacija) #######
x1 = np.array(x0)
y1 = np.array(y0)

x1 = list(x1)
for i in range(len(x1)):
    x1[i] = float(x1[i]) ** -2
x1 = np.array(x1)
x2 = np.linspace(0, max(x1) * padding, 1000)
k = coefficient(x1, y1)
y2 = k * x2


xspacing = round1(max(x1) / 8)
yspacing = round1(max(y1) / 6)
xspacing = roundPartial(xspacing, 5)
ax[1].set_xticks(np.arange(0, max(x1) * padding, xspacing))
ax[1].set_xticks(np.arange(0, max(x1) * padding, xspacing / 5), minor=True)
ax[1].set_yticks(np.arange(0, max(y1) * padding, yspacing))
ax[1].set_yticks(np.arange(0, max(y1) * padding, yspacing / 5), minor=True)
ax[1].spines['right'].set_color('#D7D7D7')
ax[1].spines['left'].set_visible(True)
ax[1].spines['top'].set_color('#D7D7D7')
ax[1].spines['bottom'].set_visible(True)
ax[1].grid(which='both')
ax[1].grid(which='minor', alpha=0.2)
ax[1].grid(which='major', alpha=0.5)


plt.subplot(2, 1, 2)
plt.plot(x1, y1, 'xk')
plt.plot(x2, y2, '-')

plt.xlim(0, max(x1) * padding)
plt.ylim(0, max(y1) * padding)
plt.tight_layout(pad=3)
plt.grid('both', 'both')
plt.title(label = r'$\bf{o(r^{-2})}$', pad = '20.0')
plt.xlabel(r'$r^{-2}[cm]$', loc = 'right')
plt.ylabel('o[lux]', loc = 'top', rotation = 0) 
plt.legend(['Meritve', r'$y = k \cdot r$'], loc ='center right')


############## PLOT 1 #################  
x1 = np.array(x0)
y1 = np.array(y0)

x2 = np.linspace(0, max(x1) * padding, 1000)
y2 = k * x2 ** -2

xspacing = round1(max(x1) / 8)
yspacing = round1(max(y1) / 6)
xspacing = roundPartial(xspacing, 5)
ax[0].set_xticks(np.arange(0, max(x1) * padding, xspacing))
ax[0].set_xticks(np.arange(0, max(x1) * padding, xspacing / 5), minor=True)
ax[0].set_yticks(np.arange(0, max(y1) * padding, yspacing))
ax[0].set_yticks(np.arange(0, max(y1) * padding, yspacing / 5), minor=True)
ax[0].spines['right'].set_color('#D7D7D7')
ax[0].spines['left'].set_visible(True)
ax[0].spines['top'].set_color('#D7D7D7')
ax[0].spines['bottom'].set_visible(True)
ax[0].grid(which='both')
ax[0].grid(which='minor', alpha=0.2)
ax[0].grid(which='major', alpha=0.5)


plt.subplot(2, 1, 1)
plt.plot(x1, y1, 'xk')
plt.plot(x2, y2, '-')

plt.xlim(0, max(x1) * padding)
plt.ylim(0, max(y1) * padding)
plt.tight_layout(pad=3)
plt.grid('both', 'both')
plt.title(label = r'$\bf{o(r)}$', pad = '20.0')
plt.xlabel(r'$r[cm]$', loc = 'right')
plt.ylabel('o[lux]', loc = 'top', rotation = 0) 
plt.legend(['Meritve', r'$y = k \cdot r^{-2}$'], loc = 'upper right')




#plt.savefig("graf.png", dpi = 600, orientation = 'portrait', papertype = 'a4', format = 'png', bbox_inches = 'tight', pad_inches = 1, transparent = None)
#plt.savefig("graf.svg", dpi = 600, orientation = 'portrait', papertype = 'a4', format = 'svg', bbox_inches = 'tight', pad_inches = 1, transparent = None)
plt.show()

